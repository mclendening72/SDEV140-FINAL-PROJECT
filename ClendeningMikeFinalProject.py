"""
Program Name: Bird-Dogey
Author:  Mike Clendening
Date written: 03/08/2025
Assignment:   Final Project for Ivy Tech SDEV140
Short Desc:   Bird-Dogey is a golf score repository program aimed at saving a golfers scores.
              The program works as follows.
                    * The user enters the golfers round information like the golfers first and last name, player ID,
                    what the golfer scored for the round, what par for the course is, and the course the golfer played
                    on.  The information is saved into an .xlsx spreadsheet.  The user can also search the spreadsheet
                    by accessing the search window and searching by the golfers playerID number.
"""

# Import all Tkinter modules
from tkinter import *
# Alias for Tkinter to use "tk." as a prefix
import tkinter as tk
# Import PhotoImage for handling images in Tkinter
from tkinter import PhotoImage
# Import image handling capabilities from PIL
from PIL import ImageTk, Image
# Import Listbox widget from tkinter
from tkinter import Listbox, END
# Import messagebox for displaying pop-ups
from tkinter import messagebox
# Import openpyxl for working with Excel files
import openpyxl
# Import os for file operations
import os, Pmw
# Import Button widget for UI interaction
from tkinter import Button
# Import tooltip library for alt-text for buttons
import tooltip
from ttkbootstrap.tooltip import ToolTip

# Global variable for the main window
main_window = None   #A global reference to the main Tkinter window

# Function to hide and show main and search windows
def open_search_screen():
    """Hide the main window and open the search screen."""
    global main_window
    if main_window:
        main_window.withdraw()  # Hide the main window
        create_search_screen()  # Open the search window

# Function to hide and show main and search windows
def restore_main_window():
    """Restore the main window when search window is closed."""
    global main_window
    search_window.destroy()
    if main_window:
        main_window.deiconify()  # Show the main window again

# Function to create the search screen
def create_splash_screen():
    """Create a splash screen that appears when the application starts."""
    splash_screen = tk.Tk()
    splash_screen.title("Bird-Dogey Splash Screen")
    splash_screen.iconbitmap('flag.ico') # Set window icon

    splash_screen.resizable(False, False) # Disable window resizing

    # Sets splash form size and centers it on any size monitor
    app_width = 611
    app_height = 422
    screen_width = splash_screen.winfo_screenwidth()
    screen_height = splash_screen.winfo_screenheight()
    x = (screen_width / 2) - (app_width / 2)
    y = (screen_height / 2) - (app_height / 2)
    splash_screen.geometry(f'{app_width}x{app_height}+{int(x)}+{int(y)}')

    try:
        image_path = PhotoImage(file=r"birddogy2.png") # Load background image
        splash_label = Label(splash_screen, image=image_path)
        splash_label.place(x=0, y=0, relwidth=1, relheight=1)
        splash_label.image = image_path  # Keep reference to prevent garbage collection
    except Exception as e:
        print(f"Error loading image: {e}")

    # Auto-close after 6 seconds
    splash_screen.after(6000, lambda: destroy_splash_and_create_main(splash_screen))
    splash_screen.mainloop()

# Function to kill the splash screen and open the main entry screen
def destroy_splash_and_create_main(splash_screen):
     """Destroy the splash screen and open the main application window."""
     splash_screen.destroy()
     create_main_window()

# Function to open search .xlsx and display the results
def search_player():
    """Retrieve and display player data from an Excel file based on PlayerID."""
    playerid = s_id_entry.get()  # Local variable holding the player's ID entered in the search window

    # Validate PlayerID (Ensure it's a 4-digit integer between 0001 and 9999)
    if not playerid:
        messagebox.showinfo("PlayerID Required", "A PlayerID is required. Please enter a PlayerID.")
        s_id_entry.focus_set()
        return

    # Ensure PlayerID is exactly 4 digits
    if not (playerid.isdigit() and len(playerid) == 4 and 1 <= int(playerid) <= 9999):
        messagebox.showerror("Invalid PlayerID", "PlayerID must be a 4-digit number between 0001 and 9999.")
        s_id_entry.delete(0, tk.END)           # Clear invalid input
        s_id_entry.focus_set()                      # Focus back on the PlayerID field
        return

    #Sets path to excel spreadsheet
    filepath = "data.xlsx"                          # Local variable holding the file path to the Excel sheet
    #Validates excel file exists / If not, create it
    if not os.path.exists(filepath):
        # Create a new workbook if the file doesn't exist
        workbook = openpyxl.Workbook()              # Local variable holding the file path to the Excel sheet
        sheet = workbook.active                     # Local variable for the active sheet of the Excel workbook
        heading = ["PlayerID", "First Name", "Last Name", "Course", "Score", "Par"]
        # Append the heading row to the sheet
        sheet.append(heading)
        # Save the workbook
        workbook.save(filepath)

    # Open the workbook
    workbook = openpyxl.load_workbook(filepath)
    # Activate the sheet in workbook
    sheet = workbook.active
    # List to store search results
    results = []                                    #Local list variable for storing search results.

    # Search through the rows of the sheet and match PlayerID
    for row in sheet.iter_rows(values_only=True):   # Local variable for iterating over rows in the Excel sheet.
        if str(row[0]) == playerid:
            results.append(row)

    # Display the search results in the UI
    if results:
        s_result_text.set("\n".join([" | ".join(map(str, r)) for r in results]))
    else:
        s_result_text.set(f"No records found for Player ID: {playerid}")

# Function to create the search screen
def create_search_screen():
    global search_window        # A reference to the search screen window
    global s_id_entry           # Entry widget for inputting the player's ID on the search screen
    global s_result_text        # String variable used to display search results on the search screen
    global s_button_img         # Image variable for the search button image on the search screen
    global s_exit_button_img    # Image variable for the exit button image on the search screen.
    global main_window          # A global reference to the main Tkinter window

    # Use Toplevel instead to create secondary screen
    search_window = tk.Toplevel(main_window)
    search_window.title("Bird-Dogey Search Screen")  # Set the title of the window
    search_window.iconbitmap('flag.ico')  # Set window icon

    search_window.resizable(False, False)  # Disable resizing of the window

    # Handle window close event to restore main_window
    search_window.protocol("WM_DELETE_WINDOW", restore_main_window)

    # Sets search form size and centers it on any size monitor
    app_width = 686
    app_height = 365
    screen_width = search_window.winfo_screenwidth()
    screen_height = search_window.winfo_screenheight()
    x = (screen_width / 2) - (app_width / 2)
    y = (screen_height / 2) - (app_height / 2)
    search_window.geometry(f'{app_width}x{app_height}+{int(x)}+{int(y)}')

    # Initialize Tooltips for the search window
    s_tooltip = Pmw.Balloon(search_window)   # A reference to the tooltip handler for the search screen

    # Try loading the background image for the search screen
    try:
        search_window.bg_image = tk.PhotoImage(file="searchbg.png")  # Keep a reference
        search_label = tk.Label(search_window, image=search_window.bg_image)
        search_label.place(x=0, y=0, relwidth=1, relheight=1)
    except tk.TclError as e:
        print(f"Error loading image: {e}")  # Handle image loading errors

    #Create Label and Textbox For ID Number
    s_id_label = Label(search_window, text="Enter Player ID:")
    s_id_label.place(x=110, y = 180)
    s_id_label.config(bg="#afed66")
    s_id_entry = Entry(search_window)  # Local reference to the player ID entry widget on the search screen
    s_id_entry.place(x=90, y=210)

    # Create Search Button
    s_button_img = Image.open("main_search_button.png")
    s_button_img = ImageTk.PhotoImage(s_button_img)  # Local reference to the search button image
    s_search = Button(search_window, image=s_button_img, bd=0, command=search_player)
    s_search.place(x=100, y=240)
    s_tooltip.bind(s_search, "Search Golfer Information")

    #Create a Label to Display Results
    s_result_text = tk.StringVar()  # Local reference to the result text variable
    s_result_label = tk.Label(search_window, textvariable=s_result_text, justify="left", wraplength=450, bg="#afed66")
    s_result_label.place(x=300, y=20)


    # Create Exit Button
    s_exit_button_img = Image.open("main_exit_button.png")  # Local reference to the exit button image
    s_exit_button_img = ImageTk.PhotoImage(s_exit_button_img)
    # A reference to the exit button on the search screen
    s_button_exit = Button(search_window, image=s_exit_button_img, bd=0, command=restore_main_window)  # A reference to the exit button on the search screen
    s_button_exit.place(x=100, y=300)
    s_tooltip.bind(s_button_exit, "Back to Main Entry")

    # Start the search window's event loop
    search_window.mainloop()

# Function to handle the data entry and validation
def enter_data(playerid_entry, firstname_entry, lastname_entry, course_listbox, score_entry, par_entry):

    # Retrieve values from entry fields and strip any extra whitespace
    playerid = playerid_entry.get().strip()         # Local variable holding the player's ID.
    firstname = firstname_entry.get().strip()       # Local variable holding the player's first name.
    lastname = lastname_entry.get().strip()         # Local variable holding the player's last name.

    # Get the selected course from the listbox
    course = course_listbox.get(tk.ANCHOR)          # Local variable holding the player's last name.

    score = score_entry.get().strip()               # Local variable holding the player's score.
    par = par_entry.get().strip()                   # Local variable holding the par for the course.

    # Validate PlayerID (Ensure it's a 4-digit integer between 0001 and 9999)
    if not playerid:
        messagebox.showinfo("PlayerID Required", "A PlayerID is required. Please enter a PlayerID.")
        playerid_entry.focus_set()  # Focus back on the PlayerID field
        return

    # Ensure PlayerID is exactly 4 digits
    if not (playerid.isdigit() and len(playerid) == 4 and 1 <= int(playerid) <= 9999):
        messagebox.showerror("Invalid PlayerID", "PlayerID must be a 4-digit number between 0001 and 9999.")
        playerid_entry.delete(0, tk.END)  # Clear invalid input
        playerid_entry.focus_set()  # Focus back on the PlayerID field
        return

    # Validate First Name (ensure it's not empty)
    if not firstname:
        messagebox.showinfo("First Name Required", "Please enter a first name.")
        firstname_entry.focus_set()  # Focus on the First Name field
        return

    # Validate Last Name (ensure it's not empty)
    if not lastname:
        messagebox.showinfo("Last Name Required", "Please enter a last name.")
        lastname_entry.focus_set()  # Focus on the Last Name field
        return

    # Validate Score (ensure not empty before conversion)
    if not score:
        messagebox.showinfo("A Score is Required", "Please enter a score.")
        score_entry.focus_set()
        return

    try:
        # Convert score to integer
        score_value = int(score)  #Local variable for converting the score into an integer.
        if not (1 <= score_value <= 200):  # Ensure score is within a valid range (0-200)
            raise ValueError  # If out of range, raise exception
    except ValueError:
        messagebox.showerror("Invalid Score", "Score must be an integer between 1 and 200.")
        score_entry.delete(0, tk.END)  # Clear invalid input
        score_entry.focus_set()  # Focus on the Score field
        return

    # Validate par (ensure not empty before conversion)
    if not par:
        messagebox.showinfo("Par for Round", "What is par for the round.")
        par_entry.focus_set()
        return

    try:
        # Convert score to integer
        par_value = int(par)             #Local variable for converting the par into an integer.
        if not (1 <= par_value <= 100):  # Ensure score is within a valid range (1-100)
            raise ValueError             # If out of range, raise exception
    except ValueError:
        messagebox.showerror("Invalid Entry", "Par must be an integer between 1 and 100.")
        par_entry.delete(0, tk.END)  # Clear invalid input
        par_entry.focus_set()  # Focus on the Score field
        return

    # Validate Course Selection
    if not course:  # Check if a course is selected
        messagebox.showinfo("Missing Course", "Please select a golf course.")
        course_listbox.focus_set()  # Focus on the course selection listbox
        return

    # Check if the Excel file exists, create it if it doesn't
    filepath = "data.xlsx"
    if not os.path.exists(filepath):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        heading = ["PlayerID", "First Name", "Last Name", "Course", "Score", "Par"]
        sheet.append(heading)  # Add headers to the sheet
        workbook.save(filepath)  # Save the workbook

    # Load the existing Excel file and append the new data
    workbook = openpyxl.load_workbook(filepath)
    sheet = workbook.active
    sheet.append([playerid, firstname, lastname, course, score, par])  # Add the new entry to the sheet
    workbook.save(filepath)  # Save the updated workbook

    # Calculate score relative to par and display the results
    score = int(score_entry.get())
    par = int(par_entry.get())
    score_round = score - par  # Local variable to hold the score relative to par

    # Determine the result based on score vs. par
    if score == par:
        create_main_window.results_label = Label(text="You shot even par today!")
        create_main_window.results_label.config(bg="#afed66", fg="red", font=("Javanese", 18))
        create_main_window.results_label.place(x=60,y=140)
    elif score < par:
        create_main_window.results_label = Label(text="You shot " + str(score_round) + " under par today!")
        create_main_window.results_label.config(bg="#afed66", fg="red", font=("Javanese", 18))
        create_main_window.results_label.place(x=60, y=140)
    elif score > par:
        create_main_window.results_label = Label(text="You shot " + str(score_round) + " over par today!")
        create_main_window.results_label.config(bg="#afed66", fg="red", font=("Javanese", 18))
        create_main_window.results_label.place(x=60, y=140)

    # Clear all fields and reset the form
    playerid_entry.delete(0, END)
    firstname_entry.delete(0, END)
    lastname_entry.delete(0, END)
    score_entry.delete(0, END)
    par_entry.delete(0, END)
    course_listbox.selection_clear(0, END)  # Deselect any selected course
    course_listbox.selection_set(0)  # Set the first course as selected
    course_listbox.activate(0)  # Activate the first item in the list

    # Set focus to PlayerID field to allow easy input of a new entry
    playerid_entry.focus_set()

# Function to create the main window (GUI)
def create_main_window():
    global main_window

    main_window = tk.Tk()  # Create the main Tkinter window
    main_window.title("Bird-Dogey Data Entry Form") # Set window title
    main_window.iconbitmap('flag.ico')  # Set window icon

    main_window.resizable(False, False)# Prevent resizing of the window

    # Set height and width for application
    app_width = 914                                     # Local variable for the width of the window.
    app_height = 675                                    # Local variable for the height of the window.
    screen_width = main_window.winfo_screenwidth()      # Local variable for the screen width.
    screen_height = main_window.winfo_screenheight()    # Local variable for the screen height.
    x = (screen_width / 2) - (app_width / 2)            # Local variable for calculating the horizontal position of the window
    y = (screen_height / 2) - (app_height / 2)          # Local variable for calculating the vertical position of the window
    main_window.geometry(f'{app_width}x{app_height}+{int(x)}+{int(y)}')

    # Initialize Pmw Tooltips
    Pmw.initialise(main_window)
    tooltip = Pmw.Balloon(main_window)  # Create a single instance for all tooltips

    # Attempt to load an image for the background (handles errors if image is missing)
    try:
        main_image_path = PhotoImage(file=r"entry_golf_guybg.png") #  Local variable for the image path to the background image of the main window
        main_label = Label(main_window, image=main_image_path) # Local variable for the label displaying the main window background image.
        main_label.place(x=0, y=0, relwidth=1, relheight=1)
        main_label.image = main_image_path  # Keep reference
    except Exception as e:
        print(f"Error loading image: {e}")

    # Create a heading/banner across Entry_Form
    #Label widget displaying the heading in the main window
    heading = Label(text="Bird-Dogey Golfer Information Form", bg="green", font="10", width="500", height="2")
    heading.pack()

    # Create labels for each entry field
    playerid_label = Label(text="Player ID: ", bg="white")      # Label widget for the player ID field
    firstname_label = Label(text="FirstName: ", bg="white")     # Label widget for the first name field
    lastname_label = Label(text="LastName: ", bg="white")       # Label widget for the last name field

    # Create listbox for course selection
    course_listbox = Listbox(main_window, font=("arial", 10), width=25, bg="#afed66")
    course_list = ["Cascades Golf Course", "Rolling Meadows", "The Point Resort", "Taylors Par 3",
                   "Bloomington Country Club", "Stone Crest", "Pine Woods", "Salt Creek Golf Course",
                   "Foxcliff Golf Course", "Other"]
    for course in course_list:
        course_listbox.insert(END, course)   # Populate the listbox with course names

    # Create other labels for score, par, and course information
    course_label = Label(text="Course Played: ", bg="white")    # Label widget for the course selection field
    score_label = Label(text="Score :", bg="white")             # Label widget for the score field
    par_label = Label(text="Par for Course :", bg="white")      # Label widget for the par field
    welcome_label = Label(text="Welcome to Bird-Dogey!", bg="#afed66", font=("Javanese", 26))  #Label widget displaying the welcome message

    # Label to display results after score entry
    result_label = tk.Label(main_window, text="", font=("Arial", 10), anchor="w", justify="left")  # Label widget displaying the results after score entry
    result_label.place(x=10, y=70)
    result_label.place_forget()

    # Places the labels and entry fields in the window at specific positions
    playerid_label.place(x=15, y=380)
    firstname_label.place(x=15, y=410)
    lastname_label.place(x=15, y=440)
    score_label.place(x=15, y=470)
    par_label.place(x=15, y=500)
    course_label.place(x=268, y=380)
    welcome_label.place(x=15, y=80)

    # Creates entry boxes for the form and places them in the window
    playerid_entry = Entry(main_window, width=6, bg="#afed66")
    firstname_entry = Entry(main_window, width=20, bg="#afed66")
    lastname_entry = Entry(main_window, width=20, bg="#afed66")
    score_entry = Entry(main_window, width=4, bg="#afed66")
    par_entry = Entry(main_window, width=4, bg="#afed66")
    playerid_entry.place(x=115, y=380)
    firstname_entry.place(x=115, y=410)
    lastname_entry.place(x=115, y=440)
    score_entry.place(x=115, y=470)
    par_entry.place(x=115, y=500)
    course_listbox.place(x=272, y=405)

    # Create buttons and their tooltips
    # Search Button
    search_button_img = Image.open("main_search_button.png")  # Image variable for the search button in the main window
    search_button_img = ImageTk.PhotoImage(search_button_img)
    button_search = Button(main_window, image=search_button_img, bd=0, command=open_search_screen)  # Button widget for the search button in the main window
    button_search.place(x=200, y=600)
    # Keep reference to image
    button_search.image = search_button_img
    # Tooltip for search button
    tooltip.bind(button_search, "Search Golfer Information")

    # Save Button
    save_button_img = Image.open("main_save_button.png")   # Image variable for the save button in the main window
    save_button_img = ImageTk.PhotoImage(save_button_img)
    # Button widget for the save button in the main window
    button_save = Button(main_window, image=save_button_img, bd=0, command=lambda: enter_data(playerid_entry,
        firstname_entry, lastname_entry, course_listbox, score_entry, par_entry))  # Local reference to the save button in the main window
    button_save.place(x=60, y=600)
    # Keep reference to image
    button_save.image = save_button_img
    # Tooltip for save button
    tooltip.bind(button_save, "Save Round Info")

    # Create Exit Button
    exit_button_img = Image.open("main_exit_button.png")  # Image variable for the exit button in the main window
    exit_button_img = ImageTk.PhotoImage(exit_button_img)
    button_exit = Button(main_window, image=exit_button_img, bd=0, command=main_window.destroy)  # Button widget for the exit button in the main window
    button_exit.place(x=340, y=600)
    # Keep reference to image
    button_exit.image = exit_button_img
    # Tooltip for exit button
    tooltip.bind(button_exit, "Exit the application")

    main_window.after(100, lambda: [main_window.lift(), main_window.focus_force(), playerid_entry.focus_set()])

    # Start the Tkinter main loop to run the GUI
    main_window.mainloop()

# Entry point to start the application
if __name__ == "__main__":
    create_splash_screen()
